#import tkinter as tk
#from tkinter import ttk
#from tkinter.messagebox import showinfo
import pygame
from openpyxl import load_workbook #import pygame_widgets
import xlsxwriter


crusherdata = load_workbook('Crushers.xlsx')
Kodiak300 = crusherdata['Kodiak300']
Jaw = crusherdata['Jaw']



#print(layout)
clock = pygame.time.Clock()
f = 0
solved = False
xst=0
xot=0
yst =0
yot = 0
xcorr=0
ycorr=0
crusher0 = [0,0]
crusher1 = [0,40]
screen1 = [20,5]
screen2 = [20,15]
screen3 = [20,25]
screen4 = [0,45]
surge1 = [0,20]
regenerate = True
running = True

#layout = {}
click = "None"
largestid = 0
largestname = 0

gradationformat = ['-1','id','pos/size',4,3.5,3,2.75,2.5,2.25,2,1.75,1.5,1.25,1,7/8,.75,5/8,.5,3/8,5/16,.25,"#4",5/32,"#8","#10","#16","#30","#40","#50","#100","#200"]
cutoff = [0,0,0,0,len(gradationformat)]


layout = [-1,0,3,400,200,1,2,1,.5,   -1,1,5,600,180,1001,     -1,2,5,600,240,1002,    -1,3,5,600,300,1003,       -1,4,6,460,100,1,   -1,5,4,320,100,2,3,    -1,6,4,200,200,1,2,
          -1,7,1,325,225,375,175,2,  -1,8,1,200,275,275,275,1,          -1,9,1,475,210,550,180,5,     -1,10,1,475,250,550,240,4,     -1,11,1,475,280,550,300,3,
          -1,12,1,475,180,590,100,6,    -1,13,1,275,175,275,250,8,      -1,14,1,400,125,350,75,7,    -2] 
connections = [-1,6,1,8,1,  -1,8,1,7,0,     -1,0,1,12,0,     -1,0,2,9,0,     -1,0,3,10,0,    -1,0,4,11,0,   -1,13,1,7,0,    -1,5,1,13,0,   -1,7,1,0,0   -1,9,1,1,0, -1,10,1,2,0,
               -1,11,1,3,0,    -1,14,1,5,0  -1,12,1,4,0,     -1,4,1,14,0]
activegradations =  [-1,100,8,100,97,93,89,85,78,70,60,49,39,29,25,21,18,14,11,9,7,5,5,4,1,1,1,1,1,1,1]
fieldgradations = []
settledgradations = []
screenhandler = []

#0 is -1 to indicate new componant
#1 Indicates part id, unique to each componant
#2 Indiacates the part. 0 is Source, 1 is Conveyor, 2 is Splitter, 3  is Screen 4 is crusher, 5 is stockpile 6 is Surge Pile, 7 is Conveyor Connection
#Crusher types, 1 is jaw 2 is kodick 200 3 is k300 4 is k400
#css in excel
while solved == False:
    for i in range(0,len(activegradations)):
        if activegradations[i] == -1:
            for j in range(0,len(layout)):
                if layout[j] == -1:
                    if layout[j+1] == activegradations[i+2]:
                        
                        if layout[j+2] == 3: #If Screen
                            
                            for k in range(0, 4): #Move to handler
                                for l in range(0,len(gradationformat)): 
                                    screenhandler = screenhandler + [activegradations[i+l]]
                            for k in range(0,len(gradationformat)): #Clear Active
                                del activegradations[i]
                            
                            cutoff = [0,0,0,0,len(gradationformat)] #Find cutoffs
                            for k in range (0, len(gradationformat)):
                                if gradationformat[k] == layout[j+6]:
                                    cutoff[1] = k
                                if gradationformat[k] == layout[j+7]:
                                    cutoff[2] = k
                                if gradationformat[k] == layout[j+8]:
                                    cutoff[3] = k

                            for k in range(0,4): #cut           #4 or 3???  
                                #screenhandler[k*len(gradationformat) + 1] = (screenhandler[k*len(gradationformat)+cutoff[k]+1] - screenhandler[k*len(gradationformat)+cutoff[k+1]-1])
                                if k == 0:
                                    screenhandler[k*len(gradationformat) + 1] = (100 - screenhandler[k*len(gradationformat)+cutoff[k+1]])  /100 *screenhandler[k*len(gradationformat) + 1]
                                if k == 1:
                                    screenhandler[k*len(gradationformat) + 1] = (screenhandler[k*len(gradationformat)+cutoff[k]] - screenhandler[k*len(gradationformat)+cutoff[k+1]])/100*screenhandler[k*len(gradationformat) + 1]
                                if k == 2:
                                    screenhandler[k*len(gradationformat) + 1] = (screenhandler[k*len(gradationformat)+cutoff[k]] - screenhandler[k*len(gradationformat)+cutoff[k+1]])/100*screenhandler[k*len(gradationformat) + 1]
                                if k == 3:
                                    screenhandler[k*len(gradationformat) + 1] = (screenhandler[k*len(gradationformat)+cutoff[k]])/100*screenhandler[k*len(gradationformat) + 1]
                                  
                                for l in range(0,len(gradationformat)):
                                    if l > 2:
                                        if l <= cutoff[k]:
                                            screenhandler[l+k*len(gradationformat)] = 100
                                        if l > cutoff[k+1]:
                                            screenhandler[l+k*len(gradationformat)] = 0
                                screenhandler[2+k*len(gradationformat)] = layout[j+1]

                            
                            for k in range(0, len(connections)): #Return to active
                                if connections[k] == -1 and connections[k+1] == layout[j+1]:
                                    if connections[k+2] == 4:
                                        screenhandler[2+3*len(gradationformat)] = connections[k+3]
                                    if connections[k+2] == 3:
                                        screenhandler[2+2*len(gradationformat)] = connections[k+3]
                                    if connections[k+2] == 2:
                                        screenhandler[2+len(gradationformat)] = connections[k+3]
                                    if connections[k+2] == 1:
                                        screenhandler[2] = connections[k+3]
                            activegradations = activegradations + screenhandler
                            
                            #for k in range(0, len(gradationformat)+1):
                                #activegradations[k+i] = 0

                            for k in range(0, len(screenhandler)): #Clear handler
                                del screenhandler[0]
                            #break

                                
                        if layout[j+2] == 1: #If Conveyor
                            for k in range(0, len(connections)):
                                if connections[k] == -1:
                                    if connections[k+1] == activegradations[i+2]:
                                        activegradations[i+2] = connections[k+3]
                                        break

                    
                        if layout[j+2] == 4: #If Crusher
                            for k in range(0, len(gradationformat)-3):
                                #print(Kodiak300.cell(row=1+k, column=9).value)
                                activegradations[i+k+3] = Kodiak300.cell(row=2+k, column=9).value
                                
                            for k in range(0, len(connections)):
                                if connections[k] == -1:
                                    if connections[k+1] == activegradations[i+2]:
                                        activegradations[i+2] = connections[k+3]
                                        break
                                    

                                    
                        if layout[j+2] == 5: #If Stock
                            for k in range (0, len(gradationformat)):
                                settledgradations = settledgradations + [activegradations[i+k]]
                                activegradations[i+k] = 0
                                #del activegradations[i]
                            #for k in range (0, len(gradationformat)):
                                #del activegradations[i+k]
                                


                            
                        if layout[j+2] == 6: #If Surge
                            for k in range(0, len(connections)):
                                if connections[k] == -1:
                                    if connections[k+1] == activegradations[i+2]:
                                        activegradations[i+2] = connections[k+3]
                                        break

    masstotal = 0                         
    for i in range(0,len(settledgradations)):
        if settledgradations[i] == -1:
            masstotal = masstotal + settledgradations[i+1]
            #print(masstotal)
            if masstotal > 99:
                solved = True
            

    #print(activegradations)
    f = f + 1
    if f == 100000:
        solved = True
    



"""
print("Step 2 cut")
print(cutoff)
print("Active", activegradations)
print("Screen", screenhandler)
"""
print("Settled", settledgradations)     



workbook = xlsxwriter.Workbook('Outputs.xlsx')
worksheet = workbook.add_worksheet("1001")


print(len(settledgradations)/len(gradationformat))



for i in range(0, int(len(settledgradations)/len(gradationformat))):
    letter = chr(ord('@')+i+1)
    for j in range(0, len(gradationformat)):
        location = str(letter) + str(j+1)
        worksheet.write(location, settledgradations[i*len(gradationformat)+j])

workbook.close()


        

Background = (0,0,0)
pygame.init()
width=1500
height=750
mdisplay=pygame.display.set_mode((width,height))
opfont= pygame.font.SysFont("candara",20)
pygame.display.set_caption("JohnFLO Hard Rock Plant")
def drawconveyor(xs,ys,xo,yo):
    pygame.draw.line(mdisplay, (255,255,255), (xs,ys), (xo,yo))
    pygame.draw.circle(mdisplay, (255,0,0), (xs, ys), 2)
    pygame.draw.circle(mdisplay, (255,0,0), (xo, yo), 2)
    menu1 = opfont.render("C"+str(layout[i+7]), 1, (255,255,255))
    mdisplay.blit(menu1,(-10+xs,5+ys))
def drawscreen(x,y):
    pygame.draw.line(mdisplay, (255,255,255), (-20+x,-5+y), (20+x,5+y))
    pygame.draw.line(mdisplay, (255,255,255), (-20+x,5+y), (20+x,15+y))
    pygame.draw.line(mdisplay, (255,255,255), (-20+x,15+y), (20+x,25+y))
    
    pygame.draw.line(mdisplay, (255,255,255), (-20+x,20+y), (-5+x,45+y))
    pygame.draw.line(mdisplay, (255,255,255), (20+x,30+y), (5+x,45+y))


    pygame.draw.circle(mdisplay, (255,0,0), (x, y), 2)
    pygame.draw.circle(mdisplay, (255,0,0), (20+x,5+y), 2)
    pygame.draw.circle(mdisplay, (255,0,0), (20+x,15+y), 2)
    pygame.draw.circle(mdisplay, (255,0,0), (20+x,25+y), 2)
    pygame.draw.circle(mdisplay, (255,0,0), (x, 45+y), 2)



    menu1 = opfont.render("Screen "+str(layout[i+5]), 1, (255,255,255))
    mdisplay.blit(menu1,(-30+x,70+y))
    menu1 = opfont.render(str(layout[i+6]), 1, (255,255,255))
    mdisplay.blit(menu1,(x-32,y))
    menu1 = opfont.render(str(layout[i+7]), 1, (255,255,255))
    mdisplay.blit(menu1,(x-32,15+y))
    menu1 = opfont.render(str(layout[i+8]), 1, (255,255,255))
    mdisplay.blit(menu1,(x-32,30+y))
def drawstockpile(x,y):
    pygame.draw.line(mdisplay, (255,255,255), (x,y), (20+x,20+y))
    pygame.draw.line(mdisplay, (255,255,255), (x,y), (-20+x,20+y))
    pygame.draw.line(mdisplay, (255,255,255), (20+x,20+y), (-20+x,20+y))
    pygame.draw.circle(mdisplay, (255,0,0), (x,y), 2)
    menu1 = opfont.render("Stockpile"+str(layout[i+1]), 1, (255,255,255))
    mdisplay.blit(menu1,(-30+x,20+y))
    menu1 = opfont.render("Product: "+str(layout[i+5]), 1, (255,255,255))
    mdisplay.blit(menu1,(-30+x,35+y))

def drawsurgepile(x,y):
    pygame.draw.line(mdisplay, (255,255,255), (x,y), (20+x,20+y))
    pygame.draw.line(mdisplay, (255,255,255), (x,y), (-20+x,20+y))
    pygame.draw.line(mdisplay, (255,255,255), (20+x,20+y), (-20+x,20+y))
    pygame.draw.line(mdisplay, (255,255,255), (x+10,y+10), (x,20+y))
    pygame.draw.line(mdisplay, (255,255,255), (x-10,y+10), (x,20+y))
    pygame.draw.circle(mdisplay, (255,0,0), (x,y), 2)
    pygame.draw.circle(mdisplay, (255,0,0), (x,20+y), 2)
    menu1 = opfont.render("Surgepile "+str(layout[i+5]), 1, (255,255,255))
    mdisplay.blit(menu1,(-30+x,30+y))

def drawsplitter(x,y):
    pygame.draw.line(mdisplay, (255,255,255), (x-20,y), (x+20,y))
    pygame.draw.line(mdisplay, (255,255,255), (x-20,y+20), (x+20,y+20))
    pygame.draw.line(mdisplay, (255,255,255), (x-20,y), (x-20,y+20))
    pygame.draw.line(mdisplay, (255,255,255), (x+20,y), (x+20,y+20))

    pygame.draw.line(mdisplay, (255,255,255), (x,y+15), (x,y+5))

    pygame.draw.circle(mdisplay, (255,0,0), (x,y), 2)
    pygame.draw.circle(mdisplay, (255,0,0), (x-20,y+10), 2)
    pygame.draw.circle(mdisplay, (255,0,0), (x+20,y+10), 2)

def drawcrusher(x,y):
    pygame.draw.line(mdisplay, (255,255,255), (x+15,y), (x+15,y+40))
    pygame.draw.line(mdisplay, (255,255,255), (x-15,y), (x-15,y+40))

    pygame.draw.line(mdisplay, (255,255,255), (x,y), (x+10,y+40))
    pygame.draw.line(mdisplay, (255,255,255), (x,y), (x-10,y+40))
    
    pygame.draw.circle(mdisplay, (255,0,0), (x+crusher0[0],y+crusher0[1]), 2)
    pygame.draw.circle(mdisplay, (255,0,0), (x+crusher1[0],y+crusher1[1]), 2)
    menu1 = opfont.render("Crusher "+str(layout[i+5]), 1, (255,255,255))
    mdisplay.blit(menu1,(-30+x,30+y))
def drawconnection(xs,ys,xo,yo):
    pygame.draw.line(mdisplay, (0,255,0), (xs,ys), (xo,yo))
    pygame.draw.line(mdisplay, (0,255,0), ((xs+xo)/2,(ys+yo)/2), ((xs+xo)/2+3,(ys+yo)/2-3))
    pygame.draw.line(mdisplay, (0,255,0), ((xs+xo)/2,(ys+yo)/2), ((xs+xo)/2-3,(ys+yo)/2-3))
    

def menu(x,y):
    pygame.draw.rect(mdisplay, (125, 125, 125), (x,y,50,50))
    pygame.draw.line(mdisplay, (0,0,255), (x,y), (x+50,y))
    pygame.draw.line(mdisplay, (0,0,255), (x,y+50), (x+50,y+50))
    pygame.draw.line(mdisplay, (0,0,255), (x,y), (x,y+50))
    pygame.draw.line(mdisplay, (0,0,255), (x+50,y), (x+50,y+50))
    


def newcomponant(partid,parttype,x,y):
    print("e")


def generate():
    print("test")




while running == True:
    #pygame.event.get()




    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            running = False
            
        if event.type == pygame.MOUSEBUTTONDOWN:
            if event.button == 1:  # Check for left mouse button click
                x, y = pygame.mouse.get_pos()
                print(f"Clicked at ({x}, {y})")

                if click == "CPlace":
                    for i in range (0, len(layout)):
                        if layout[i] == -1:
                            if layout[i+1] > largestid:
                                largestid = layout[i+1]
                            if layout[i+2] > largestname:
                                largestname = layout[i+2]
                        if layout[i] == -2 and click=="CPlace":
                            click = "None"
                            layout[i] = -1
                            layout = layout + [largestid + 1, 1, x, y, x+40, y, largestname + 1, -2]
                            regenerate = True

                if click == "SplitterPlace":
                    for i in range (0, len(layout)):
                        if layout[i] == -1:
                            if layout[i+1] > largestid:
                                largestid = layout[i+1]
                            if layout[i+2] > largestname:
                                largestname = layout[i+2]
                        if layout[i] == -2 and click=="SplitterPlace":
                            click= "None"
                            layout[i] =  -1
                            layout = layout + [largestid + 1, 2, x, y, largestname + 1, -2]
                            regenerate = True
                            
                            
                if click == "ScreenPlace":
                    for i in range (0, len(layout)):
                        if layout[i] == -1:
                            if layout[i+1] > largestid:
                                largestid = layout[i+1]
                            if layout[i+2] > largestname:
                                largestname = layout[i+2]
                        if layout[i] == -2 and click=="ScreenPlace":
                            click= "None"
                            layout[i] =  -1
                            layout = layout + [largestid + 1, 3, x, y, largestname + 1, 0, 0, 0, -2]
                            regenerate = True
                            
                            
                if click == "CrusherPlace":
                    for i in range (0, len(layout)):
                        if layout[i] == -1:
                            if layout[i+1] > largestid:
                                largestid = layout[i+1]
                            if layout[i+2] > largestname:
                                largestname = layout[i+2]
                        if layout[i] == -2 and click=="CrusherPlace":
                            click= "None"
                            layout[i] =  -1
                            layout = layout + [largestid + 1, 4, x, y, largestname + 1, -2]
                            regenerate = True                        

                if click == "StockpilePlace":
                    for i in range (0, len(layout)):
                        if layout[i] == -1:
                            if layout[i+1] > largestid:
                                largestid = layout[i+1]
                            if layout[i+2] > largestname:
                                largestname = layout[i+2]
                        if layout[i] == -2 and click=="StockpilePlace":
                            click= "None"
                            layout[i] =  -1
                            layout = layout + [largestid + 1, 5, x, y, 1000, -2]
                            regenerate = True

                if click == "SurgepilePlace":
                    for i in range (0, len(layout)):
                        if layout[i] == -1:
                            if layout[i+1] > largestid:
                                largestid = layout[i+1]
                            if layout[i+2] > largestname:
                                largestname = layout[i+2]
                        if layout[i] == -2 and click=="SurgepilePlace":
                            click= "None"
                            layout[i] =  -1
                            layout = layout + [largestid + 1, 6, x, y, largestname + 1, -2]
                            regenerate = True

                if y < 60 and x > 70 and x < 120:
                    click = "CPlace"
                if y < 60 and x > 130 and x < 180:
                    click = "SplitterPlace"
                if y < 60 and x > 190 and x < 240:
                    click = "ScreenPlace"
                if y < 60 and x > 250 and x < 300:
                    click = "CrusherPlace"
                if y < 60 and x > 310 and x < 360:
                    click = "StockpilePlace"
                if y < 60 and x > 370 and x < 420:
                    click = "SurgepilePlace"
                    
                for i in range (0, len(layout)):
                    if layout[i] == -1:
                        if layout[i+3] - x < 10 and layout[i+3] - x > -10 and layout[i+4] - y < 10 and layout[i+4] - y > -10:
                            print(layout[i+1])




    
    while regenerate == True:

    
        mdisplay.fill(Background)

        menu(10,10)


        
        menu(70,10)
        pygame.draw.line(mdisplay, (255,255,255), (80,50), (115,25))

        
        menu(130,10)
        """
        pygame.draw.line(mdisplay, (255,255,255), (200,15), (240,15))
        pygame.draw.line(mdisplay, (255,255,255), (200,35), (240,35))
        pygame.draw.line(mdisplay, (255,255,255), (200,15), (200,35))
        pygame.draw.line(mdisplay, (255,255,255), (240,15), (240,35))
        pygame.draw.line(mdisplay, (255,255,255), (220,30), (220,20))
        """
        menu(190,10)

        pygame.draw.line(mdisplay, (255,255,255), (195,40), (235,50))
        pygame.draw.line(mdisplay, (255,255,255), (195,30), (235,40))
        pygame.draw.line(mdisplay, (255,255,255), (195,20), (235,30))
        
        menu(250,10)
        pygame.draw.line(mdisplay, (255,255,255), (255,15), (255,55))
        pygame.draw.line(mdisplay, (255,255,255), (285,15), (285,55))
        pygame.draw.line(mdisplay, (255,255,255), (260,55), (270,15))
        pygame.draw.line(mdisplay, (255,255,255), (280,55), (270,15))

        menu(310,10)
        pygame.draw.line(mdisplay, (255,255,255), (335,25), (355,45))
        pygame.draw.line(mdisplay, (255,255,255), (335,25), (315,45))
        pygame.draw.line(mdisplay, (255,255,255), (315,45), (355,45))
        menu(370,10)
        pygame.draw.line(mdisplay, (255,255,255), (395,25), (415,45))
        pygame.draw.line(mdisplay, (255,255,255), (395,25), (375,45))
        pygame.draw.line(mdisplay, (255,255,255), (375,45), (415,45))
        pygame.draw.line(mdisplay, (255,255,255), (385,35), (395,45))
        pygame.draw.line(mdisplay, (255,255,255), (405,35), (395,45))


        
        
        for i in range (0,len(layout)):
            if layout[i] == -1:
                if layout[i+2] == 1:
                    drawconveyor(layout[i+3],layout[i+4],layout[i+5],layout[i+6])
                if layout[i+2] == 2:
                    drawsplitter(layout[i+3],layout[i+4])
                if layout[i+2] == 3:
                    drawscreen(layout[i+3],layout[i+4])
                if layout[i+2] == 4:
                    drawcrusher(layout[i+3],layout[i+4])
                if layout[i+2] == 5:
                    drawstockpile(layout[i+3],layout[i+4])
                if layout[i+2] == 6:
                    drawsurgepile(layout[i+3],layout[i+4])

        for i in range (0,len(connections)):
            if connections[i] == -1:
                for j in range(0,len(layout)):
                    if layout[j] == -1 and connections[i+1] == layout[j+1]:
                        adjustmentx=0
                        adjustmenty=0
                        if layout[j+2] == 1 and connections[i+2] == 1:
                            adjustmentx = layout[j+5]-layout[j+3]
                            adjustmenty = layout[j+6]-layout[j+4]
                            
                        if layout[j+2] == 3:
                            if connections[i+2] == 1:
                                adjustmentx = screen1[0]
                                adjustmenty = screen1[1]
                            if connections[i+2] == 2:
                                adjustmentx = screen2[0]
                                adjustmenty = screen2[1]
                            if connections[i+2] == 3:
                                adjustmentx = screen3[0]
                                adjustmenty = screen3[1]
                            if connections[i+2] == 4:
                                adjustmentx = screen4[0]
                                adjustmenty = screen4[1]

                                
                        if layout[j+2] == 4:
                            adjustmentx = crusher1[0]
                            adjustmenty = crusher1[1]
                        if layout[j+2] == 6:
                            adjustmentx = surge1[0]
                            adjustmenty = surge1[1]

                            
                        xs=layout[j+3] + adjustmentx
                        ys=layout[j+4] + adjustmenty
                    if layout[j] == -1 and connections[i+3] == layout[j+1]:
                        xo=layout[j+3]
                        yo=layout[j+4]
                drawconnection(xs,ys,xo,yo)

                
            

            regenerate = False
        pygame.display.update()

pygame.quit()
sys.exit()




